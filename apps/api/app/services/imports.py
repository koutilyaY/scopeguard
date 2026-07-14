"""CSV/XLSX import parsing, mapping, validation, preview and commit.

Each importer validates rows individually and reports errors by row number, so a
single bad row never silently corrupts an import.
"""

import csv
import hashlib
import io
import uuid
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Invoice, InvoiceLine, Project, TimeEntry, WorkItem
from app.models.enums import BillableStatus, InvoiceStatus, WorkItemStatus
from app.services.review.duplicates import time_entry_content_hash

MAX_IMPORT_ROWS = 5000
MAX_DAILY_MINUTES = 16 * 60  # excessively large daily entry threshold
KNOWN_ROLES_HINT = 200  # roles longer than this are suspicious


@dataclass
class RowError:
    row: int
    field: str
    message: str


@dataclass
class ImportPreview:
    columns: list[str]
    sample_rows: list[dict]
    total_rows: int
    valid_rows: int
    errors: list[RowError] = field(default_factory=list)
    warnings: list[RowError] = field(default_factory=list)


@dataclass
class ImportResult:
    created: int
    skipped_duplicates: int
    errors: list[RowError] = field(default_factory=list)
    warnings: list[RowError] = field(default_factory=list)


def parse_tabular(filename: str, data: bytes) -> tuple[list[str], list[dict]]:
    """Parse CSV or XLSX into (columns, rows-as-dicts). Values become strings."""
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return [], []
        rows = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            row = {}
            for idx, col in enumerate(header):
                value = values[idx] if idx < len(values) else None
                if isinstance(value, datetime):
                    row[col] = value.date().isoformat()
                elif isinstance(value, date):
                    row[col] = value.isoformat()
                else:
                    row[col] = str(value).strip() if value is not None else ""
            rows.append(row)
            if len(rows) > MAX_IMPORT_ROWS:
                raise ValueError(f"Import exceeds {MAX_IMPORT_ROWS} rows")
        workbook.close()
        return header, rows

    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    columns = [c.strip() for c in (reader.fieldnames or [])]
    rows = []
    for raw in reader:
        rows.append({(k or "").strip(): (v or "").strip() for k, v in raw.items()})
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValueError(f"Import exceeds {MAX_IMPORT_ROWS} rows")
    return columns, rows


def _parse_date(value: str) -> date | None:
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _mapped(row: dict, mapping: dict[str, str], field_name: str) -> str:
    column = mapping.get(field_name, "")
    return row.get(column, "").strip() if column else ""


# --------------------------------------------------------------------- work items

JIRA_FIELDS = [
    "external_id",
    "title",
    "description",
    "status",
    "work_type",
    "assignee",
    "created_at",
    "completed_at",
    "source_url",
]

JIRA_STATUS_MAP = {
    "to do": WorkItemStatus.open,
    "open": WorkItemStatus.open,
    "backlog": WorkItemStatus.open,
    "in progress": WorkItemStatus.in_progress,
    "in review": WorkItemStatus.in_progress,
    "done": WorkItemStatus.done,
    "closed": WorkItemStatus.done,
    "resolved": WorkItemStatus.done,
    "cancelled": WorkItemStatus.cancelled,
    "canceled": WorkItemStatus.cancelled,
}


def validate_work_item_rows(
    rows: list[dict], mapping: dict[str, str]
) -> tuple[list[dict], list[RowError], list[RowError]]:
    valid, errors, warnings = [], [], []
    seen_ids: set[str] = set()
    for index, row in enumerate(rows, start=2):  # row 1 is the header
        external_id = _mapped(row, mapping, "external_id")
        title = _mapped(row, mapping, "title")
        if not external_id:
            errors.append(RowError(index, "external_id", "Missing work item ID"))
            continue
        if not title:
            errors.append(RowError(index, "title", "Missing title"))
            continue
        if external_id in seen_ids:
            warnings.append(
                RowError(index, "external_id", f"Duplicate ID {external_id} in file — skipped")
            )
            continue
        seen_ids.add(external_id)
        status_raw = _mapped(row, mapping, "status").lower()
        status = JIRA_STATUS_MAP.get(status_raw)
        if status is None:
            warnings.append(
                RowError(index, "status", f"Unknown status '{status_raw}' — defaulting to open")
            )
            status = WorkItemStatus.open
        created_at = _parse_date(_mapped(row, mapping, "created_at"))
        completed_at = _parse_date(_mapped(row, mapping, "completed_at"))
        valid.append(
            {
                "external_id": external_id,
                "title": title,
                "description": _mapped(row, mapping, "description") or None,
                "status": status,
                "work_type": _mapped(row, mapping, "work_type") or None,
                "assignee": _mapped(row, mapping, "assignee") or None,
                "created_at_external": datetime.combine(created_at, datetime.min.time(), UTC)
                if created_at
                else None,
                "completed_at_external": datetime.combine(completed_at, datetime.min.time(), UTC)
                if completed_at
                else None,
                "source_url": _mapped(row, mapping, "source_url") or None,
                "raw_payload": row,
            }
        )
    return valid, errors, warnings


def commit_work_items(
    db: Session, organization_id: uuid.UUID, project: Project, valid_rows: list[dict]
) -> ImportResult:
    existing_ids = {
        r
        for r in db.execute(
            select(WorkItem.external_id).where(
                WorkItem.organization_id == organization_id,
                WorkItem.project_id == project.id,
            )
        ).scalars()
    }
    created = skipped = 0
    for row in valid_rows:
        if row["external_id"] in existing_ids:
            skipped += 1
            continue
        content_hash = hashlib.sha256(
            f"{row['external_id']}|{row['title']}|{row['status'].value}".encode()
        ).hexdigest()
        db.add(
            WorkItem(
                organization_id=organization_id,
                project_id=project.id,
                external_system="jira_csv",
                content_hash=content_hash,
                **row,
            )
        )
        created += 1
    db.commit()
    return ImportResult(created=created, skipped_duplicates=skipped)


# ------------------------------------------------------------------- time entries

TIMESHEET_FIELDS = [
    "employee_name",
    "employee_role",
    "work_date",
    "hours",
    "minutes",
    "description",
    "billable_status",
    "work_item_external_id",
    "external_id",
]


def validate_time_entry_rows(
    rows: list[dict], mapping: dict[str, str]
) -> tuple[list[dict], list[RowError], list[RowError]]:
    valid, errors, warnings = [], [], []
    daily_totals: dict[tuple[str, date], int] = {}
    for index, row in enumerate(rows, start=2):
        employee = _mapped(row, mapping, "employee_name")
        if not employee:
            errors.append(RowError(index, "employee_name", "Missing employee name"))
            continue
        work_date = _parse_date(_mapped(row, mapping, "work_date"))
        if work_date is None:
            errors.append(RowError(index, "work_date", "Missing or unparseable date"))
            continue
        if work_date.year < 2000 or work_date > date.today().replace(year=date.today().year + 1):
            errors.append(RowError(index, "work_date", f"Impossible date {work_date}"))
            continue

        minutes: int | None = None
        minutes_raw = _mapped(row, mapping, "minutes")
        hours_raw = _mapped(row, mapping, "hours")
        try:
            if minutes_raw:
                minutes = int(Decimal(minutes_raw))
            elif hours_raw:
                minutes = int(Decimal(hours_raw) * 60)
        except (InvalidOperation, ValueError):
            errors.append(
                RowError(index, "hours", f"Unparseable time value '{minutes_raw or hours_raw}'")
            )
            continue
        if minutes is None:
            errors.append(RowError(index, "hours", "Missing hours/minutes"))
            continue
        if minutes <= 0:
            errors.append(RowError(index, "hours", f"Non-positive time ({minutes} minutes)"))
            continue

        key = (employee.lower(), work_date)
        daily_totals[key] = daily_totals.get(key, 0) + minutes
        if daily_totals[key] > MAX_DAILY_MINUTES:
            warnings.append(
                RowError(
                    index,
                    "hours",
                    f"{employee} exceeds {MAX_DAILY_MINUTES / 60:.0f}h on {work_date} — verify",
                )
            )

        billable_raw = _mapped(row, mapping, "billable_status").lower()
        if billable_raw in ("yes", "true", "billable", "1"):
            billable = BillableStatus.billable
        elif billable_raw in ("no", "false", "non-billable", "nonbillable", "0"):
            billable = BillableStatus.non_billable
        else:
            billable = BillableStatus.unknown

        role = _mapped(row, mapping, "employee_role") or None
        if role and len(role) > KNOWN_ROLES_HINT:
            errors.append(RowError(index, "employee_role", "Role value too long"))
            continue

        valid.append(
            {
                "employee_name": employee,
                "employee_role": role,
                "work_date": work_date,
                "minutes": minutes,
                "description": _mapped(row, mapping, "description") or None,
                "billable_status": billable,
                "work_item_external_id": _mapped(row, mapping, "work_item_external_id") or None,
                "external_id": _mapped(row, mapping, "external_id") or None,
            }
        )
    return valid, errors, warnings


def commit_time_entries(
    db: Session, organization_id: uuid.UUID, project: Project, valid_rows: list[dict]
) -> ImportResult:
    work_items = {
        w.external_id: w.id
        for w in db.execute(
            select(WorkItem).where(
                WorkItem.organization_id == organization_id, WorkItem.project_id == project.id
            )
        ).scalars()
    }
    existing_hashes = {
        h
        for h in db.execute(
            select(TimeEntry.content_hash).where(
                TimeEntry.organization_id == organization_id,
                TimeEntry.project_id == project.id,
            )
        ).scalars()
    }
    created = skipped = 0
    warnings: list[RowError] = []
    for row_index, row in enumerate(valid_rows, start=2):
        content_hash = time_entry_content_hash(
            str(project.id),
            row["employee_name"],
            str(row["work_date"]),
            row["minutes"],
            row["description"],
        )
        if content_hash in existing_hashes:
            skipped += 1
            continue
        work_item_ref = row.pop("work_item_external_id")
        work_item_id = work_items.get(work_item_ref) if work_item_ref else None
        if work_item_ref and work_item_id is None:
            warnings.append(
                RowError(
                    row_index,
                    "work_item",
                    f"Unknown work item '{work_item_ref}' — entry imported unlinked",
                )
            )
        db.add(
            TimeEntry(
                organization_id=organization_id,
                project_id=project.id,
                work_item_id=work_item_id,
                source="csv_import",
                content_hash=content_hash,
                **row,
            )
        )
        # NOTE: intra-file duplicates are intentionally imported (flagged later by
        # the review engine's duplicate detection), matching real-world timesheets.
        created += 1
    db.commit()
    return ImportResult(created=created, skipped_duplicates=skipped, warnings=warnings)


# ---------------------------------------------------------------------- invoices

INVOICE_FIELDS = [
    "invoice_number",
    "billing_period_start",
    "billing_period_end",
    "issue_date",
    "status",
    "currency",
    "line_description",
    "service_category",
    "quantity",
    "unit_price",
    "amount",
    "work_item_external_id",
]


def validate_invoice_rows(
    rows: list[dict], mapping: dict[str, str]
) -> tuple[list[dict], list[RowError], list[RowError]]:
    valid, errors, warnings = [], [], []
    for index, row in enumerate(rows, start=2):
        number = _mapped(row, mapping, "invoice_number")
        if not number:
            errors.append(RowError(index, "invoice_number", "Missing invoice number"))
            continue
        try:
            amount_raw = _mapped(row, mapping, "amount") or "0"
            amount_minor = int(
                (Decimal(amount_raw.replace("$", "").replace(",", "")) * 100).to_integral_value()
            )
            unit_raw = _mapped(row, mapping, "unit_price") or "0"
            unit_minor = int(
                (Decimal(unit_raw.replace("$", "").replace(",", "")) * 100).to_integral_value()
            )
            quantity = Decimal(_mapped(row, mapping, "quantity") or "1")
        except InvalidOperation:
            errors.append(RowError(index, "amount", "Unparseable monetary value"))
            continue
        status_raw = (_mapped(row, mapping, "status") or "issued").lower()
        try:
            status = InvoiceStatus(status_raw)
        except ValueError:
            warnings.append(
                RowError(index, "status", f"Unknown status '{status_raw}' — defaulting to issued")
            )
            status = InvoiceStatus.issued
        valid.append(
            {
                "invoice_number": number,
                "billing_period_start": _parse_date(_mapped(row, mapping, "billing_period_start")),
                "billing_period_end": _parse_date(_mapped(row, mapping, "billing_period_end")),
                "issue_date": _parse_date(_mapped(row, mapping, "issue_date")),
                "status": status,
                "currency": (_mapped(row, mapping, "currency") or "USD").upper()[:3],
                "line_description": _mapped(row, mapping, "line_description") or "(imported line)",
                "service_category": _mapped(row, mapping, "service_category") or None,
                "quantity": quantity,
                "unit_price_minor": unit_minor,
                "amount_minor": amount_minor,
                "work_item_external_id": _mapped(row, mapping, "work_item_external_id") or None,
            }
        )
    return valid, errors, warnings


def commit_invoices(
    db: Session, organization_id: uuid.UUID, project: Project, valid_rows: list[dict]
) -> ImportResult:
    """Rows sharing an invoice_number become lines of one invoice."""
    work_items = {
        w.external_id: w.id
        for w in db.execute(
            select(WorkItem).where(
                WorkItem.organization_id == organization_id, WorkItem.project_id == project.id
            )
        ).scalars()
    }
    existing_numbers = {
        n
        for n in db.execute(
            select(Invoice.invoice_number).where(Invoice.organization_id == organization_id)
        ).scalars()
    }
    grouped: dict[str, list[dict]] = {}
    for row in valid_rows:
        grouped.setdefault(row["invoice_number"], []).append(row)

    created = skipped = 0
    for number, rows in grouped.items():
        if number in existing_numbers:
            skipped += 1
            continue
        head = rows[0]
        subtotal = sum(r["amount_minor"] for r in rows)
        invoice = Invoice(
            organization_id=organization_id,
            project_id=project.id,
            invoice_number=number,
            billing_period_start=head["billing_period_start"],
            billing_period_end=head["billing_period_end"],
            issue_date=head["issue_date"],
            currency=head["currency"],
            status=head["status"],
            subtotal_minor=subtotal,
            tax_minor=0,
            total_minor=subtotal,
        )
        db.add(invoice)
        db.flush()
        for row in rows:
            db.add(
                InvoiceLine(
                    organization_id=organization_id,
                    invoice_id=invoice.id,
                    description=row["line_description"],
                    service_category=row["service_category"],
                    quantity=row["quantity"],
                    unit_price_minor=row["unit_price_minor"],
                    amount_minor=row["amount_minor"],
                    linked_work_item_id=work_items.get(row["work_item_external_id"])
                    if row["work_item_external_id"]
                    else None,
                )
            )
        created += 1
    db.commit()
    return ImportResult(created=created, skipped_duplicates=skipped)
